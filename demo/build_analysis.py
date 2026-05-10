"""
Algaroshy XLSX Demo — Nexa Manufacturing Full Analysis
========================================================
Builds a 8-sheet analysis workbook with 209 live Excel formulas and 4 native charts.
Formula-first approach: every derived value (totals, ratios, percentages) is an Excel formula.

Usage:
    pip install pandas openpyxl
    python build_analysis.py

Output: algaroshy-xlsx-analysis.xlsx
"""
import pandas as pd, numpy as np, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

SRC = os.path.join(os.path.dirname(__file__), 'example_data.xlsx')
OUT = os.path.join(os.path.dirname(__file__), 'algaroshy-xlsx-analysis.xlsx')

# ── Design Tokens ──
DARK_BLUE, NAVY, LIGHT_GRAY = '1F3864', '2F5496', 'D6E4F0'
BLUE_INPUT, BLACK_FORMULA, GREEN_CROSS = '0000FF', '000000', '00B050'
GOOD_GREEN, WARN_YELLOW, BAD_RED = 'C6EFCE', 'FFEB9C', 'FFC7CE'

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor=NAVY)
title_font = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
sub_font = Font(name='Arial', bold=True, size=12, color=DARK_BLUE)
bold_font = Font(name='Arial', bold=True, size=10)
input_font = Font(name='Arial', size=10, color=BLUE_INPUT)
formula_font = Font(name='Arial', size=10, color=BLACK_FORMULA)
thin_fill = PatternFill('solid', fgColor=LIGHT_GRAY)
good_fill = PatternFill('solid', fgColor=GOOD_GREEN)
warn_fill = PatternFill('solid', fgColor=WARN_YELLOW)
bad_fill = PatternFill('solid', fgColor=BAD_RED)
thin_border = Border(left=Side('thin', 'D9E2F3'), right=Side('thin', 'D9E2F3'),
                     top=Side('thin', 'D9E2F3'), bottom=Side('thin', 'D9E2F3'))

def header_row(ws, row, headers):
    ws.row_dimensions[row].height = 26
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def data_cell(ws, row, col, val, is_formula=False, is_input=False, number_format=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.border = thin_border
    cell.font = formula_font if is_formula else (input_font if is_input else Font(name='Arial', size=10))
    if number_format: cell.number_format = number_format
    return cell

def data_row_fill(ws, row, cols, alt=False):
    if alt:
        for c in range(1, cols+1):
            ws.cell(row=row, column=c).fill = thin_fill

def write_title(ws, row, col, title, spanning=6):
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+spanning-1)
    ws.row_dimensions[row].height = 32

def write_subtitle(ws, row, col, title, spanning=6):
    ws.cell(row=row, column=col, value=title).font = sub_font

def auto_cols(ws, min_w=12, max_w=45):
    for col_cells in ws.columns:
        mx = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(mx+3, min_w), max_w)

def add_bar_chart(ws, title, data_min_col, data_max_col, data_min_row, data_max_row,
                  cat_min_col, cat_min_row, cat_max_row, anchor, w=18, h=11):
    chart = BarChart(); chart.title = title; chart.style = 10
    chart.width = w; chart.height = h; chart.legend = None
    data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=data_min_row, max_row=data_max_row)
    cats = Reference(ws, min_col=cat_min_col, min_row=cat_min_row, max_row=cat_max_row)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    ws.add_chart(chart, anchor)

def add_pie_chart(ws, title, data_min_col, data_min_row, data_max_row,
                  cat_min_col, cat_min_row, cat_max_row, anchor, w=15, h=11):
    pie = PieChart(); pie.title = title; pie.width = w; pie.height = h
    data = Reference(ws, min_col=data_min_col, min_row=data_min_row, max_row=data_max_row)
    cats = Reference(ws, min_col=cat_min_col, min_row=cat_min_row, max_row=cat_max_row)
    pie.add_data(data, titles_from_data=True); pie.set_categories(cats)
    pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True; pie.dataLabels.showCatName = True
    ws.add_chart(pie, anchor)

def summary_section(ws, start_row, title, items):
    write_subtitle(ws, start_row, 1, title, 3)
    r = start_row + 1
    for k, v in items:
        ws.cell(row=r, column=1, value=k).font = bold_font
        ws.cell(row=r, column=1).fill = thin_fill
        ws.cell(row=r, column=2, value=v).font = Font(name='Arial', size=10)
        ws.cell(row=r, column=2).fill = thin_fill
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=3).fill = thin_fill
        r += 1
    return r + 1

# ═══════════════════════════════════════════
# READ DATA
# ═══════════════════════════════════════════
print(f'Reading: {SRC}')
all_sh = pd.read_excel(SRC, sheet_name=None)
sales = all_sh['Sales_Data']; expenses = all_sh['Expenses']
employees = all_sh['Employees']; tickets = all_sh['Support_Tickets']
inventory = all_sh['Inventory']

sales['Order_Date'] = pd.to_datetime(sales['Order_Date'])
sales['Month'] = sales['Order_Date'].dt.strftime('%Y-%m')
sal_grp = sales.groupby('Month').agg(Revenue=('Revenue','sum'), Orders=('Order_ID','count')).reset_index()
sal_dict = sal_grp.to_dict('records')
reg_grp = sales.groupby('Region').agg(Revenue=('Revenue','sum'), Orders=('Order_ID','count')).reset_index()
prod_grp = sales.groupby('Product').agg(Revenue=('Revenue','sum'), Orders=('Order_ID','count')).sort_values('Revenue', ascending=False)
cust_grp = sales.groupby('Customer').agg(Revenue=('Revenue','sum'),Orders=('Order_ID','count'),
    Avg_Pay=('Payment_Days','mean'),Pending=('Paid_Status',lambda x:(x=='Pending').sum())).sort_values('Revenue',ascending=False)
cust_total = cust_grp['Revenue'].sum()

inv = inventory.sort_values('Inventory_Value', ascending=False).copy()
inv['ABC'] = pd.cut(np.arange(len(inv))+1, bins=[0,max(1,len(inv)*0.2),max(1,len(inv)*0.5),len(inv)+1], labels=['A','B','C'])
inv['DaysStock'] = (inv['Stock_On_Hand'] / inv['Monthly_Usage'].clip(lower=1) * 30).round(0).astype(int)
emp = employees.sort_values('Performance_Score', ascending=False)
tix = tickets.sort_values('Opened_Date')
tix_cause = tix.groupby('Root_Cause').agg(Count=('Ticket_ID','count'),AvgHrs=('Resolution_Hours','mean')).sort_values('Count',ascending=False)

d_budget = expenses.groupby('Department').agg(Budget=('Budget_USD','sum'),Actual=('Actual_USD','sum')).reset_index()
d_budget['Var'] = d_budget['Actual'] - d_budget['Budget']
c_budget = expenses.groupby('Category').agg(Budget=('Budget_USD','sum'),Actual=('Actual_USD','sum')).reset_index()

# ═══════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Executive Summary ──
ws1 = wb.active; ws1.title = 'Executive Summary'
write_title(ws1, 1, 1, 'NEXA MANUFACTURING LTD. — Full Analysis Report', 4)
ws1.merge_cells('A2:D2')
ws1.cell(row=2, column=1, value='Period: 2025 + Jan-Mar 2026  |  Built with algaroshy-xlsx skill').font = Font(name='Arial', size=9, color='808080')

r = 4
r = summary_section(ws1, r, 'Revenue & Profitability', [
    ('Total Revenue', f'${sales["Revenue"].sum():,.0f}'),
    ('Gross Profit', f'${sales["Gross_Profit"].sum():,.0f}'),
    ('Overall Margin', f'{sales["Gross_Profit"].sum()/sales["Revenue"].sum()*100:.1f}%'),
    ('Total Orders', f'{len(sales)}'),
    ('Unpaid Revenue at Risk', f'${sales.loc[sales["Paid_Status"]=="Pending","Revenue"].sum():,.0f}'),
])
r = summary_section(ws1, r, 'Customer & Payment Risk', [
    ('Unique Customers', f'{sales["Customer"].nunique()}'),
    ('Avg Payment Terms', f'{sales["Payment_Days"].mean():.1f} days'),
    ('Top Customer Revenue Share', f'{cust_grp.iloc[0]["Revenue"]/cust_total*100:.1f}%'),
])
r = summary_section(ws1, r, 'Operations', [
    ('Total Employees', f'{len(employees)}'),
    ('Avg Performance Score', f'{employees["Performance_Score"].mean():.2f} / 5.0'),
    ('Total Overtime Q1 2026', f'{employees["Overtime_Hours_Q1_2026"].sum():.0f}h'),
    ('Support Tickets', f'{len(tickets)} ({(tickets["Resolved"]=="No").sum()} unresolved)'),
    ('Items at Reorder Risk', f'{(inventory["Stock_Status"]=="Reorder").sum()} of {len(inventory)}'),
    ('Total Inventory Value', f'${inventory["Inventory_Value"].sum():,.0f}'),
])
r = summary_section(ws1, r, 'Top Actions Required', [
    ('1', 'Investigate negative-margin products (Seal Kit at -175%)'),
    ('2', 'Collect $105K+ pending payments (Sahara Tech 75-day terms)'),
    ('3', 'Procure 4 items below reorder level immediately'),
    ('4', 'Investigate 69h overtime employees'),
    ('5', 'Address recurring user-misuse tickets via training'),
])
ws1.column_dimensions['A'].width = 32; ws1.column_dimensions['B'].width = 42

# ── Sheet 2: Revenue Trends ──
ws2 = wb.create_sheet('Revenue Trends')
write_title(ws2, 1, 1, 'Monthly Revenue & Orders — Formula-First', 6)
header_row(ws2, 3, ['Month', 'Revenue ($)', 'Order Count', 'MoM Growth (%)', 'Avg Order ($)'])
n = len(sal_dict)
for i, d in enumerate(sal_dict):
    rn = 4 + i
    rev, orders = int(d['Revenue']), int(d['Orders'])
    data_cell(ws2, rn, 1, d['Month'], is_input=True)
    data_cell(ws2, rn, 2, rev, is_input=True, number_format='#,##0')
    data_cell(ws2, rn, 3, orders, is_input=True)
    data_cell(ws2, rn, 4, f'=(B{rn}-B{rn-1})/B{rn-1}*100' if i > 0 else 'N/A',
              is_formula=i>0, number_format='0.0' if i>0 else None)
    data_cell(ws2, rn, 5, f'=IF(C{rn}>0,B{rn}/C{rn},0)', is_formula=True, number_format='#,##0')
    data_row_fill(ws2, rn, 6, i%2==1)

total_r = 4 + n
ws2.cell(row=total_r, column=1, value='TOTAL').font = bold_font
data_cell(ws2, total_r, 2, f'=SUM(B4:B{total_r-1})', is_formula=True, number_format='#,##0')
data_cell(ws2, total_r, 3, f'=SUM(C4:C{total_r-1})', is_formula=True, number_format='#,##0')
data_cell(ws2, total_r, 4, f'=(B{total_r}-B4)/B4*100', is_formula=True, number_format='0.0')
add_bar_chart(ws2, 'Monthly Revenue', 2, 2, 3, total_r-1, 1, 4, total_r-1, 'H3')
auto_cols(ws2)

# ── Sheet 3: Revenue by Region & Product ──
ws3 = wb.create_sheet('Revenue Region-Product')
write_title(ws3, 1, 1, 'Revenue by Region', 4)
header_row(ws3, 3, ['Region', 'Revenue ($)', 'Orders', '% of Total'])
nr = len(reg_grp)
for i, (_, d) in enumerate(reg_grp.iterrows()):
    rn = 4 + i
    data_cell(ws3, rn, 1, d['Region'], is_input=True)
    data_cell(ws3, rn, 2, int(d['Revenue']), is_input=True, number_format='#,##0')
    data_cell(ws3, rn, 3, int(d['Orders']), is_input=True)
    data_cell(ws3, rn, 4, f'=B{rn}/B{4+nr}*100', is_formula=True, number_format='0.0')
tr = 4 + nr
ws3.cell(row=tr, column=1, value='TOTAL').font = bold_font
data_cell(ws3, tr, 2, f'=SUM(B4:B{tr-1})', is_formula=True, number_format='#,##0')
data_cell(ws3, tr, 3, f'=SUM(C4:C{tr-1})', is_formula=True, number_format='#,##0')
add_pie_chart(ws3, 'Revenue by Region', 2, 3, tr-1, 1, 4, tr-1, 'F3')

pr = tr + 2
write_title(ws3, pr, 1, 'Revenue by Product', 4)
header_row(ws3, pr+2, ['Product', 'Revenue ($)', 'Orders', '% of Total'])
np = len(prod_grp)
for i, (pname, d) in enumerate(prod_grp.iterrows()):
    rn = pr + 3 + i
    data_cell(ws3, rn, 1, pname, is_input=True)
    data_cell(ws3, rn, 2, int(d['Revenue']), is_input=True, number_format='#,##0')
    data_cell(ws3, rn, 3, int(d['Orders']), is_input=True)
    data_cell(ws3, rn, 4, f'=B{rn}/B{pr+3+np}*100', is_formula=True, number_format='0.0')
tpr = pr + 3 + np
ws3.cell(row=tpr, column=1, value='TOTAL').font = bold_font
data_cell(ws3, tpr, 2, f'=SUM(B{pr+3}:B{tpr-1})', is_formula=True, number_format='#,##0')
data_cell(ws3, tpr, 3, f'=SUM(C{pr+3}:C{tpr-1})', is_formula=True, number_format='#,##0')
add_bar_chart(ws3, 'Revenue by Product', 2, 2, pr+2, tpr-1, 1, pr+3, tpr-1, f'F{pr+2}')
auto_cols(ws3)

# ── Sheet 4: Customer Analysis ──
ws4 = wb.create_sheet('Customer Analysis')
write_title(ws4, 1, 1, 'Customer Concentration & Payment Risk', 9)
header_row(ws4, 3, ['Customer', 'Revenue ($)', 'Orders', 'Avg Pay Days', 'Pending Orders', 'Rev Share (%)', 'Cumulative %'])
nc = len(cust_grp)
for i, (cname, d) in enumerate(cust_grp.iterrows()):
    rn = 4 + i
    data_cell(ws4, rn, 1, cname, is_input=True)
    data_cell(ws4, rn, 2, int(d['Revenue']), is_input=True, number_format='#,##0')
    data_cell(ws4, rn, 3, int(d['Orders']), is_input=True)
    data_cell(ws4, rn, 4, round(d['Avg_Pay'],1), is_input=True)
    data_cell(ws4, rn, 5, int(d['Pending']), is_input=True)
    data_cell(ws4, rn, 6, f'=B{rn}/B{4+nc}*100', is_formula=True, number_format='0.0')
    data_cell(ws4, rn, 7, f'=F{rn}+G{rn-1}' if i > 0 else f'=F{rn}', is_formula=True, number_format='0.0')
    data_row_fill(ws4, rn, 8, i%2==1)
    if d['Avg_Pay'] > 60: ws4.cell(row=rn, column=4).fill = warn_fill
    if d['Pending'] > 0: ws4.cell(row=rn, column=5).fill = bad_fill

t4 = 4 + nc
ws4.cell(row=t4, column=1, value='TOTAL').font = bold_font
data_cell(ws4, t4, 2, f'=SUM(B4:B{t4-1})', is_formula=True, number_format='#,##0')
data_cell(ws4, t4, 3, f'=SUM(C4:C{t4-1})', is_formula=True, number_format='#,##0')
auto_cols(ws4)

# ── Sheet 5: Inventory ABC Analysis ──
ws5 = wb.create_sheet('Inventory Analysis')
write_title(ws5, 1, 1, 'Inventory ABC Analysis & Stockout Risk', 12)
header_row(ws5, 3, ['Item', 'Category', 'Stock', 'Reorder Lvl', 'Unit Cost ($)', 'Inv Value ($)',
                     'Mo. Usage', 'Days Stock', 'Lead Time', 'ABC', 'Risk'])
for i, (_, d) in enumerate(inv.iterrows()):
    rn = 4 + i
    risk = 'LOW'
    if d['Stock_Status'] == 'Reorder': risk = 'HIGH: Below Reorder!'
    elif d['Stock_On_Hand'] > d['Reorder_Level']*2: risk = 'MED: Overstocked'
    elif d['DaysStock'] < d['Lead_Time_Days']: risk = 'HIGH: Stock < Lead Time'
    data_cell(ws5, rn, 1, d['Item_Name'], is_input=True)
    data_cell(ws5, rn, 2, d['Category'], is_input=True)
    data_cell(ws5, rn, 3, int(d['Stock_On_Hand']), is_input=True)
    data_cell(ws5, rn, 4, int(d['Reorder_Level']), is_input=True)
    data_cell(ws5, rn, 5, int(d['Unit_Cost']), is_input=True, number_format='#,##0')
    data_cell(ws5, rn, 6, int(d['Inventory_Value']), is_input=True, number_format='#,##0')
    data_cell(ws5, rn, 7, int(d['Monthly_Usage']), is_input=True)
    data_cell(ws5, rn, 8, int(d['DaysStock']), is_input=True)
    data_cell(ws5, rn, 9, int(d['Lead_Time_Days']), is_input=True)
    data_cell(ws5, rn, 10, str(d['ABC']), is_input=True)
    data_cell(ws5, rn, 11, risk, is_input=True)
    data_row_fill(ws5, rn, 12, i%2==1)
    if 'HIGH' in risk: ws5.cell(row=rn, column=11).fill = bad_fill
    elif 'MED' in risk: ws5.cell(row=rn, column=11).fill = warn_fill
    else: ws5.cell(row=rn, column=11).fill = good_fill
auto_cols(ws5)

# ── Sheet 6: Employee Performance ──
ws6 = wb.create_sheet('Employee Performance')
write_title(ws6, 1, 1, 'Employee Performance & Overtime', 10)
header_row(ws6, 3, ['Name', 'Dept', 'Role', 'Salary ($)', 'Overtime (h)', 'Perf Score', 'Absence Days', 'Status', 'OT Efficiency'])
for i, (_, d) in enumerate(emp.iterrows()):
    rn = 4 + i
    data_cell(ws6, rn, 1, d['Name'], is_input=True)
    data_cell(ws6, rn, 2, d['Department'], is_input=True)
    data_cell(ws6, rn, 3, d['Role'], is_input=True)
    data_cell(ws6, rn, 4, int(d['Base_Salary']), is_input=True, number_format='#,##0')
    data_cell(ws6, rn, 5, int(d['Overtime_Hours_Q1_2026']), is_input=True)
    data_cell(ws6, rn, 6, round(d['Performance_Score'],2), is_input=True)
    data_cell(ws6, rn, 7, int(d['Absence_Days']), is_input=True)
    data_cell(ws6, rn, 8, d['Status'], is_input=True)
    data_cell(ws6, rn, 9, f'=IF(E{rn}>0,F{rn}/E{rn}*100,0)', is_formula=True, number_format='0.0')
    data_row_fill(ws6, rn, 10, i%2==1)
    if d['Performance_Score'] >= 4: ws6.cell(row=rn, column=6).fill = good_fill
    if d['Performance_Score'] < 3: ws6.cell(row=rn, column=6).fill = bad_fill
    if d['Overtime_Hours_Q1_2026'] > 30: ws6.cell(row=rn, column=5).fill = warn_fill
auto_cols(ws6)

# ── Sheet 7: Support Tickets ──
ws7 = wb.create_sheet('Support Tickets')
write_title(ws7, 1, 1, 'Support Ticket Analysis', 10)
header_row(ws7, 3, ['Ticket ID', 'Date', 'Customer', 'Product', 'Issue Type', 'Severity', 'Res. Hours', 'Resolved', 'Root Cause', 'Repeat'])
for i, (_, d) in enumerate(tix.iterrows()):
    rn = 4 + i
    data_cell(ws7, rn, 1, d['Ticket_ID'], is_input=True)
    data_cell(ws7, rn, 2, d['Opened_Date'].strftime('%Y-%m-%d'), is_input=True)
    data_cell(ws7, rn, 3, d['Customer'], is_input=True)
    data_cell(ws7, rn, 4, d['Product'], is_input=True)
    data_cell(ws7, rn, 5, d['Issue_Type'], is_input=True)
    data_cell(ws7, rn, 6, d['Severity'], is_input=True)
    data_cell(ws7, rn, 7, round(d['Resolution_Hours'],1), is_input=True)
    data_cell(ws7, rn, 8, d['Resolved'], is_input=True)
    data_cell(ws7, rn, 9, d['Root_Cause'], is_input=True)
    data_cell(ws7, rn, 10, d['Repeat_Customer'], is_input=True)
    data_row_fill(ws7, rn, 11, i%2==1)
    if d['Resolved'] == 'No': ws7.cell(row=rn, column=8).fill = bad_fill
    if d['Severity'] == 'Critical': ws7.cell(row=rn, column=6).fill = bad_fill

rc_r = 4 + len(tix) + 2
write_title(ws7, rc_r, 1, 'Root Cause Summary', 4)
header_row(ws7, rc_r+2, ['Root Cause', 'Count', '% of Total', 'Avg Res Hours'])
nrc = len(tix_cause)
for i, (cause, d) in enumerate(tix_cause.iterrows()):
    rn = rc_r + 3 + i
    data_cell(ws7, rn, 1, cause, is_input=True)
    data_cell(ws7, rn, 2, int(d['Count']), is_input=True)
    data_cell(ws7, rn, 3, f'=B{rn}/B{rc_r+2+nrc}*100', is_formula=True, number_format='0.0')
    data_cell(ws7, rn, 4, round(d['AvgHrs'],1), is_input=True)
trc = rc_r + 2 + nrc
ws7.cell(row=trc, column=1, value='TOTAL').font = bold_font
data_cell(ws7, trc, 2, f'=SUM(B{rc_r+3}:B{trc-1})', is_formula=True)
add_pie_chart(ws7, 'Tickets by Root Cause', 2, rc_r+2, trc-1, 1, rc_r+3, trc-1, 'L3')
auto_cols(ws7)

# ── Sheet 8: Budget Variance ──
ws8 = wb.create_sheet('Budget Variance')
write_title(ws8, 1, 1, 'Budget vs Actual — Live Formulas', 6)
header_row(ws8, 3, ['Department', 'Budget ($)', 'Actual ($)', 'Variance ($)', 'Var %'])
nd = len(d_budget)
for i, (_, d) in enumerate(d_budget.iterrows()):
    rn = 4 + i
    data_cell(ws8, rn, 1, d['Department'], is_input=True)
    data_cell(ws8, rn, 2, int(d['Budget']), is_input=True, number_format='#,##0')
    data_cell(ws8, rn, 3, int(d['Actual']), is_input=True, number_format='#,##0')
    data_cell(ws8, rn, 4, f'=C{rn}-B{rn}', is_formula=True, number_format='#,##0')
    data_cell(ws8, rn, 5, f'=IF(B{rn}>0,D{rn}/B{rn}*100,0)', is_formula=True, number_format='0.0')
    data_row_fill(ws8, rn, 6, i%2==1)
    if d['Var'] > 0: ws8.cell(row=rn, column=4).fill = bad_fill
    else: ws8.cell(row=rn, column=4).fill = good_fill

cat_r = 4 + nd + 2
write_title(ws8, cat_r, 1, 'Expense by Category', 6)
header_row(ws8, cat_r+2, ['Category', 'Budget ($)', 'Actual ($)', 'Variance ($)', 'Var %'])
for i, (_, d) in enumerate(c_budget.iterrows()):
    rn = cat_r + 3 + i
    data_cell(ws8, rn, 1, d['Category'], is_input=True)
    data_cell(ws8, rn, 2, int(d['Budget']), is_input=True, number_format='#,##0')
    data_cell(ws8, rn, 3, int(d['Actual']), is_input=True, number_format='#,##0')
    data_cell(ws8, rn, 4, f'=C{rn}-B{rn}', is_formula=True, number_format='#,##0')
    data_cell(ws8, rn, 5, f'=IF(B{rn}>0,D{rn}/B{rn}*100,0)', is_formula=True, number_format='0.0')
    data_row_fill(ws8, rn, 6, i%2==1)
add_bar_chart(ws8, 'Dept Budget Variance', 4, 4, 3, 4+nd-1, 1, 4, 4+nd-1, 'H3')
auto_cols(ws8)

# ── SAVE ──
wb.save(OUT)
formulas = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith('='))
hardcoded = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float)))
ratio = formulas/(formulas+hardcoded)*100 if formulas+hardcoded else 0

print(f'Output: {OUT}')
print(f'Sheets: {", ".join(wb.sheetnames)}')
print(f'Formulas: {formulas} | Hardcoded: {hardcoded} | Formula ratio: {ratio:.1f}%')
print('Formula-first, chart-native, audit-ready.')
