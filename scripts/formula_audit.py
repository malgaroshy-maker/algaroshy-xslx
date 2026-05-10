"""Audit formula vs hardcode ratio in an xlsx file."""
import sys, os, json
from openpyxl import load_workbook

def audit(filepath, json_output=False):
    wb = load_workbook(filepath, data_only=False)
    result = {'sheets': {}, 'total_formulas': 0, 'total_hardcoded': 0, 'total_cells': 0}
    
    for ws in wb.worksheets:
        formulas = 0
        hardcoded = 0
        cross_sheet = 0
        strings = 0
        empty = 0
        
        for row in ws.iter_rows():
            for cell in row:
                result['total_cells'] += 1
                v = cell.value
                if v is None:
                    empty += 1
                elif isinstance(v, str) and v.startswith('='):
                    formulas += 1
                    if '!' in v:
                        cross_sheet += 1
                elif isinstance(v, (int, float)):
                    hardcoded += 1
                else:
                    strings += 1
        
        total = formulas + hardcoded
        ratio = formulas / total * 100 if total > 0 else 0
        result['sheets'][ws.title] = {
            'formulas': formulas,
            'hardcoded': hardcoded,
            'cross_sheet_formulas': cross_sheet,
            'strings': strings,
            'empty': empty,
            'formula_ratio': round(ratio, 1)
        }
        result['total_formulas'] += formulas
        result['total_hardcoded'] += hardcoded
    
    total = result['total_formulas'] + result['total_hardcoded']
    result['overall_ratio'] = round(result['total_formulas'] / total * 100, 1) if total > 0 else 0
    
    if json_output:
        print(json.dumps(result, indent=2))
    else:
        print(f'{"Sheet":<30} {"Formulas":>8} {"Hardcoded":>10} {"Ratio":>8} {"Cross-Sheet":>12}')
        print('-' * 72)
        for name, s in result['sheets'].items():
            print(f'{name:<30} {s["formulas"]:>8} {s["hardcoded"]:>10} {s["formula_ratio"]:>7.1f}% {s["cross_sheet_formulas"]:>12}')
        print('-' * 72)
        print(f'{"TOTAL":<30} {result["total_formulas"]:>8} {result["total_hardcoded"]:>10} {result["overall_ratio"]:>7.1f}%')
    
    return result

if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(description='Audit formula vs hardcode ratio')
    ap.add_argument('file')
    ap.add_argument('--json', action='store_true')
    args = ap.parse_args()
    audit(args.file, args.json)
