"""Apply financial color standards to openpyxl workbook output."""
import sys, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
GREEN_FONT = Font(name='Arial', size=10, color='00B050')
BLACK_FONT = Font(name='Arial', size=10, color='000000')
YELLOW_FILL = PatternFill('solid', fgColor='FFFF00')

GOOD_FILL = PatternFill('solid', fgColor='C6EFCE')
WARN_FILL = PatternFill('solid', fgColor='FFEB9C')
BAD_FILL = PatternFill('solid', fgColor='FFC7CE')

def is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith('=')

def is_cross_sheet(formula_text):
    return '!' in formula_text

def is_numeric_input(cell):
    return isinstance(cell.value, (int, float)) and not isinstance(cell.value, str)

def has_conditional_fill(ws, cell):
    """Check if cell already has a conditional fill applied."""
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = str(cell.fill.fgColor.rgb)
        if rgb in ('00C6EFCE', '00FFEB9C', '00FFC7CE'):
            return True
    return False

def apply_financial_colors(filename, output=None):
    wb = load_workbook(filename)
    changed = 0
    
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if has_conditional_fill(ws, cell):
                    continue
                if is_formula(cell):
                    if is_cross_sheet(cell.value):
                        cell.font = GREEN_FONT
                        changed += 1
                    else:
                        cell.font = BLACK_FONT
                        changed += 1
                elif is_numeric_input(cell):
                    cell.font = BLUE_FONT
                    changed += 1
    
    out = output or filename
    wb.save(out)
    return changed

def apply_conditional_fill(filename, column_map, output=None):
    """
    Apply green/yellow/red fills based on column_map rules.
    column_map: dict of column_letter -> {'good': pattern, 'warn': pattern, 'bad': pattern}
    """
    wb = load_workbook(filename)
    changed = 0
    
    for ws_name, rules in column_map.items():
        if ws_name not in wb.sheetnames:
            continue
        ws = wb[ws_name]
        for col_letter, rule in rules.items():
            col_idx = ord(col_letter.upper()) - ord('A') + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val is None:
                    continue
                if isinstance(val, str) and val.startswith('='):
                    continue
                try:
                    v = float(val)
                    if v > 0:
                        cell.fill = GOOD_FILL if 'good' in rule else BAD_FILL
                    elif v < 0:
                        cell.fill = BAD_FILL
                    changed += 1
                except (ValueError, TypeError):
                    pass
    
    out = output or filename
    wb.save(out)
    return changed

if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('file')
    ap.add_argument('--output', '-o')
    ap.add_argument('--financial-colors', action='store_true', help='Apply blue/black/green financial colors')
    ap.add_argument('--audit', action='store_true', help='Report formula vs hardcode counts')
    args = ap.parse_args()
    
    if args.financial_colors:
        n = apply_financial_colors(args.file, args.output)
        print(f'Applied financial colors to {n} cells')
    else:
        wb = load_workbook(args.file)
        formulas = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row if is_formula(c))
        hardcoded = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row if is_numeric_input(c))
        total = formulas + hardcoded
        ratio = formulas / total * 100 if total else 0
        print(f'Formulas: {formulas}, Hardcoded: {hardcoded}, Formula ratio: {ratio:.1f}%')
